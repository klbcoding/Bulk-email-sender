import openpyxl
import smtplib
import sys
from email_templates import rejection_email_template


# Establishing SMTP connection
server = smtplib.SMTP("smtp.gmail.com", 587)
# Enable Transport Layer Security (TLS)
server.starttls()

# Configuring details
sender = "sender@gmail.com"
app_password = "abcd efgh ijkl mnop"

# Logging into account
try:
    server.login(sender, app_password)
except smtplib.SMTPAuthenticationError:
    print("Error signing in, check that the login details are correct")
    sys.exit()
else:
    print("Login successful.")

# Loading Excel data
wb = openpyxl.load_workbook("rejected_applicants.xlsx")
ws = wb["Sheet1"]

# Running program
index = 2
emails_sent = 0
while True:
    name = ws[f"A{index}"].value
    recipient = ws[f"B{index}"].value
    position = ws[f"C{index}"].value

    # Check if the last value is reached by checking for None
    if name is None:
        break

    subject = f"Job Application Outcome for {position}"
    message = f"""From: {sender}
To: {recipient}
Subject: {subject}\n
{rejection_email_template(name, position)}
"""

    server.sendmail(sender, recipient, message)
    emails_sent += 1
    print(f"{emails_sent} email(s) sent!")
    index += 1


